"""
excel_report.py — Build a two-sheet scored Excel workbook from portfolio evaluation results.

Sheet 1 "Scored Results": One row per result.  Input fields colour-coded by Review Status
(green = CONFIRMED, amber = AUTO_POPULATED, red = NOT_FOUND/FLAGGED).  Score columns
colour-coded green ≥4.0, amber 2.5–3.9, red <2.5.

Sheet 2 "Summary": Portfolio-level gap analysis, weakest dimensions, methodology footnote.

No Streamlit dependency — pure openpyxl.  Returns an in-memory bytes object
suitable for st.download_button or an API response.
"""
from __future__ import annotations

import io
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False


# ---------------------------------------------------------------------------
# Column-level help notes — shown as hover comments on header cells
# ---------------------------------------------------------------------------
_COLUMN_NOTES: dict[str, str] = {
    "Confidence (0–5)": (
        "CONFIDENCE — how much should we trust the evidence?\n"
        "Score >=4.0 = Submission-ready (green)\n"
        "2.5–3.9 = Needs improvement (amber)\n"
        "<2.5 = High risk (red)\n\n"
        "Made up of:\n"
        "  Directness  — how directly does evidence link to the result?\n"
        "  Verification — was evidence independently reviewed?\n"
        "  Recency     — how current is the evidence?"
    ),
    "Clarity (0–5)": (
        "CLARITY — can someone else interpret this result the same way?\n"
        "Score >=4.0 = Submission-ready (green)\n"
        "2.5–3.9 = Needs improvement (amber)\n"
        "<2.5 = High risk (red)\n\n"
        "Made up of:\n"
        "  Definition   — who/what/where/when specified?\n"
        "  Measurement  — collection method and sampling disclosed?\n"
        "  Integrity    — complete data with audit trail?\n"
        "  Scope        — coverage matches the claim?\n"
        "  Governance   — named owner and decision use stated?"
    ),
    "Verdict": (
        "VERDICT — combined judgement across both axes:\n"
        "  'Strong KPI — submission-ready on both axes' = both >=4.0\n"
        "  'Misleading KPI'  = strong evidence, unclear claim\n"
        "  'Well-defined but weak evidence' = clear claim, weak evidence\n"
        "  'High risk'       = both axes need significant work"
    ),
    "% of Target": (
        "% OF TARGET = Actual Achievement / Approved Target * 100\n"
        "Calculated automatically from the logframe columns.\n"
        "Blank if either target or achievement was not found in the document."
    ),
    "Direction Mismatch": (
        "DIRECTION MISMATCH:\n"
        "  'Yes' = the indicator implies a change in one direction\n"
        "  (e.g. 'increase in employment') but the baseline-to-achievement\n"
        "  comparison shows the OPPOSITE direction.\n"
        "  Donors will flag this — review the result framing before submitting."
    ),
    "Fix 1": "TOP PRIORITY FIX — the single action that would most improve this result's score. Address this before submitting.",
    "Fix 2": "SECOND PRIORITY FIX — address after Fix 1.",
    "Fix 3": "THIRD PRIORITY FIX — address after Fixes 1 and 2.",
    "Review Status": (
        "REVIEW STATUS — did ImpactProof find this field with high confidence?\n\n"
        "  CONFIRMED (green)       = extracted with high confidence.\n"
        "                           Still verify before sharing with donors.\n\n"
        "  AUTO_POPULATED (amber)  = AI extracted, medium confidence.\n"
        "                           REVIEW REQUIRED — check against source document.\n\n"
        "  NOT_FOUND (red)         = field was absent from the document.\n"
        "                           Fill manually before sharing.\n\n"
        "Do NOT share this Excel with donors until all amber and red cells\n"
        "have been reviewed and confirmed."
    ),
}

# ---------------------------------------------------------------------------
# Colour palette (hex, no leading #)
# ---------------------------------------------------------------------------
_GREEN_DARK   = "1B5E20"   # header / strong score
_GREEN_LIGHT  = "C8E6C9"   # confirmed field / score ≥ 4.0
_AMBER_DARK   = "8A6500"   # header text on amber
_AMBER_LIGHT  = "FFF9C4"   # auto-populated field / score 2.5–3.9
_RED_DARK     = "B71C1C"   # header text on red
_RED_LIGHT    = "FFCDD2"   # flagged/not-found / score < 2.5
_GREY_LIGHT   = "F5F5F5"   # alternating row
_WHITE        = "FFFFFF"
_HEADER_BG    = "1B5E20"   # dark green header
_HEADER_FG    = "FFFFFF"
_SCORE_GREEN  = "C8E6C9"
_SCORE_AMBER  = "FFF9C4"
_SCORE_RED    = "FFCDD2"

# Review status values
STATUS_CONFIRMED      = "CONFIRMED"
STATUS_AUTO_POPULATED = "AUTO_POPULATED"
STATUS_NOT_FOUND      = "NOT_FOUND"
STATUS_FLAGGED        = "FLAGGED"

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

# (key_in_df, display_header, width)
_INPUT_COLUMNS = [
    ("indicator_name",       "Indicator Name",             30),
    ("result_statement",     "Result Statement",           40),
    ("target_group",         "Target Group",               22),
    ("timeframe",            "Timeframe",                  16),
    ("geographic_scope",     "Geographic Scope",           22),
    ("evidence_type",        "Evidence Type",              20),
    ("evidence_description", "Evidence Description",       35),
    ("evidence_date",        "Evidence Date",              14),
    ("internal_review",      "Internal Review",            20),
    ("external_review",      "External Review",            20),
    ("verifier",             "Verifier",                   22),
    ("logframe_indicator",   "Logframe Indicator",         30),
    ("logframe_baseline",    "Baseline Value",             18),
    ("logframe_target",      "Approved Target",            18),
    ("logframe_achievement", "Actual Achievement",         18),
    ("learning_notes",       "Learning / Adaptation",      25),
    ("limitations_notes",    "Limitations",                25),
    ("beneficiary_voice",    "Beneficiary Voice",          24),
    ("additional_context",   "Result Owner & Decision",    25),
    ("sector",               "Sector",                     16),
    ("primary_donor",        "Primary Donor",              16),
]

_SCORE_COLUMNS = [
    ("confidence_score",   "Confidence (0–5)",    12),
    ("clarity_score",      "Clarity (0–5)",       12),
    ("verdict",            "Verdict",             30),
    ("pct_of_target",      "% of Target",         12),
    ("direction_mismatch", "Direction Mismatch",  16),
    ("fix_1",              "Fix 1",               30),
    ("fix_2",              "Fix 2",               30),
    ("fix_3",              "Fix 3",               30),
    ("review_status",      "Review Status",       18),
]


def _fill(hex_colour: str):
    return PatternFill("solid", fgColor=hex_colour)


def _font(bold=False, colour="000000", size=10):
    return Font(bold=bold, color=colour, size=size)


def _border_thin():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


def _status_fill(status: str) -> PatternFill:
    if status == STATUS_CONFIRMED:
        return _fill(_GREEN_LIGHT)
    if status == STATUS_AUTO_POPULATED:
        return _fill(_AMBER_LIGHT)
    return _fill(_RED_LIGHT)


def _score_fill(score: float | None) -> PatternFill:
    if score is None:
        return _fill(_WHITE)
    if score >= 4.0:
        return _fill(_SCORE_GREEN)
    if score >= 2.5:
        return _fill(_SCORE_AMBER)
    return _fill(_SCORE_RED)


def build_scored_excel(
    rows: list[dict],
    evaluations: list[dict],
    field_statuses: list[dict] | None = None,
    org_name: str = "",
    document_name: str = "",
) -> bytes:
    """Build a two-sheet Excel workbook and return it as bytes.

    Parameters
    ----------
    rows : list[dict]
        One dict per result, keys matching _INPUT_COLUMNS key names
        (i.e., the portfolio submission dicts).
    evaluations : list[dict]
        Parallel list of evaluation dicts from evaluator.evaluate_submission().
    field_statuses : list[dict] | None
        Parallel list of {field_key: STATUS_*} dicts.  If None, all fields
        are treated as AUTO_POPULATED.
    org_name, document_name : str
        Optional metadata for the Summary sheet header.
    """
    if not _OPENPYXL_OK:
        raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")

    wb = Workbook()
    _build_sheet1(wb, rows, evaluations, field_statuses)
    _build_sheet2(wb, rows, evaluations, org_name, document_name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Sheet 1 — Scored Results
# ---------------------------------------------------------------------------

def _build_sheet1(wb: "Workbook", rows, evaluations, field_statuses):
    ws = wb.active
    ws.title = "Scored Results"
    ws.freeze_panes = "A2"  # freeze header row

    all_cols = _INPUT_COLUMNS + _SCORE_COLUMNS
    n_input  = len(_INPUT_COLUMNS)

    # --- Header row ---
    for col_idx, (key, header, width) in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill    = _fill(_HEADER_BG)
        cell.font    = _font(bold=True, colour=_HEADER_FG, size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border  = _border_thin()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        # Add hover note for score columns that have guidance text
        if _OPENPYXL_OK and header in _COLUMN_NOTES:
            note = Comment(_COLUMN_NOTES[header], author="ImpactProof")
            note.width  = 340
            note.height = 150
            cell.comment = note

    ws.row_dimensions[1].height = 28

    # --- Data rows ---
    for row_idx, (sub, ev) in enumerate(zip(rows, evaluations), start=2):
        statuses = (field_statuses or [{}] * len(rows))[row_idx - 2]
        is_alt   = (row_idx % 2 == 0)
        base_bg  = _GREY_LIGHT if is_alt else _WHITE

        # Input columns
        for col_idx, (key, _, _) in enumerate(_INPUT_COLUMNS, start=1):
            value  = sub.get(key, "")
            status = statuses.get(key, STATUS_AUTO_POPULATED)
            cell   = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.fill      = _status_fill(status)
            cell.font      = _font(size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border    = _border_thin()

        # Score columns
        fixes    = ev.get("fixes", [])
        linkage  = ev.get("logframe_linkage", {})
        pct      = linkage.get("pct_of_target")
        dir_miss = linkage.get("direction_mismatch", False)
        conf     = ev.get("confidence_score")
        clar     = ev.get("clarity_score")
        verdict  = ev.get("verdict", "")
        overall_status = statuses.get("_overall", STATUS_AUTO_POPULATED)

        score_values = [
            (conf,    "confidence_score"),
            (clar,    "clarity_score"),
            (verdict, "verdict"),
            (f"{pct:.0f}%" if pct is not None else "—", "pct_of_target"),
            ("Yes" if dir_miss else "No",                "direction_mismatch"),
            (fixes[0]["message"] if len(fixes) > 0 else "", "fix_1"),
            (fixes[1]["message"] if len(fixes) > 1 else "", "fix_2"),
            (fixes[2]["message"] if len(fixes) > 2 else "", "fix_3"),
            (overall_status,                               "review_status"),
        ]

        for local_idx, (value, key) in enumerate(score_values):
            col_idx = n_input + local_idx + 1
            cell    = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = _font(bold=(key in ("confidence_score", "clarity_score")), size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top",
                                       horizontal="center" if key in ("confidence_score", "clarity_score",
                                                                       "pct_of_target", "direction_mismatch")
                                       else "left")
            cell.border    = _border_thin()

            # Score colour coding
            if key == "confidence_score":
                cell.fill = _score_fill(conf)
            elif key == "clarity_score":
                cell.fill = _score_fill(clar)
            elif key == "direction_mismatch" and dir_miss:
                cell.fill = _fill(_RED_LIGHT)
                cell.font = _font(bold=True, colour=_RED_DARK, size=9)
            elif key == "review_status":
                cell.fill = _status_fill(overall_status)
            else:
                cell.fill = _fill(base_bg)

        ws.row_dimensions[row_idx].height = 48

    # Legend row at bottom
    leg_row = len(rows) + 3
    ws.cell(row=leg_row, column=1,
            value="Legend: Field colours = Review Status").font = _font(bold=True, size=9)
    for leg_col, (label, colour) in enumerate([
        ("CONFIRMED (user-verified)", _GREEN_LIGHT),
        ("AUTO-POPULATED (extracted by AI — review before submitting)", _AMBER_LIGHT),
        ("NOT FOUND / FLAGGED", _RED_LIGHT),
    ], start=2):
        c = ws.cell(row=leg_row, column=leg_col, value=label)
        c.fill = _fill(colour)
        c.font = _font(size=9)
        c.border = _border_thin()


# ---------------------------------------------------------------------------
# Sheet 2 — Summary
# ---------------------------------------------------------------------------

def _build_sheet2(wb, rows, evaluations, org_name, document_name):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 42

    timestamp = datetime.now().strftime("%d %b %Y %H:%M")
    n = len(evaluations)

    # Title — Council XXVI: Portfolio Decision Audit framing
    ws.merge_cells("A1:C1")
    title = ws["A1"]
    title.value = "ImpactProof — Portfolio Decision Audit"
    title.fill  = _fill(_HEADER_BG)
    title.font  = _font(bold=True, colour=_HEADER_FG, size=14)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Metadata — value cells merged B:C so long text wraps in full width (cols B+C = ~64 chars)
    _meta_rows = [
        ("Organisation",    org_name or "(not specified)"),
        ("Document",        document_name or "(not specified)"),
        ("Results scored",  n),          # int, not str — avoids "number stored as text" warning
        ("Generated",       timestamp),
        ("Methodology",     "USAID ADS 201 · FCDO 2025 · Bond Evidence Principles 2024 · World Bank RF · OECD-DAC 2019"),
        ("Scoring engine",  "ImpactProof — deterministic decision-making engine. "
                            "Rule-based scoring anchored to named donor standards. "
                            "No AI judgement applied to scores — same inputs always produce the same determination."),
        ("Disclaimer",      "Scores reflect evidence patterns in the submitted document. "
                            "Each determination is traceable to a named sub-criterion "
                            "(USAID ADS 201.3.5.7 for Validity, Integrity, Precision, Reliability, Timeliness). "
                            "This is a pre-submission quality check, not a donor audit. "
                            "All auto-populated fields (amber) must be reviewed before treating scores as final. "
                            "Red cells = not found in document — fill manually before sharing with donors."),
    ]
    _meta_heights = [20, 20, 16, 16, 28, 48, 72]
    for idx, ((label, value), row_h) in enumerate(zip(_meta_rows, _meta_heights)):
        row = idx + 3
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font      = _font(bold=True, size=10)
        lbl.alignment = Alignment(wrap_text=True, vertical="top")
        # Merge B and C so the value has full width to wrap within
        ws.merge_cells(f"B{row}:C{row}")
        cell = ws.cell(row=row, column=2, value=value)
        cell.font          = _font(size=10)
        cell.alignment     = Alignment(wrap_text=True, vertical="top")
        if isinstance(value, int):
            cell.number_format = "0"   # display as plain integer, suppress text-storage warning
        ws.row_dimensions[row].height = row_h

    # Portfolio determinations table — Council XXVI framing
    start_row = 12
    ws.cell(row=start_row, column=1,
            value="PORTFOLIO DETERMINATIONS").font = _font(bold=True, colour=_GREEN_DARK, size=11)
    start_row += 1

    headers = ["Indicator / Result", "Confidence", "Clarity", "Determination", "% of Target", "Direction OK?"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=col, value=h)
        c.fill = _fill(_HEADER_BG)
        c.font = _font(bold=True, colour=_HEADER_FG, size=10)
        c.border = _border_thin()
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[start_row].height = 20
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    for i, (sub, ev) in enumerate(zip(rows, evaluations)):
        r = start_row + 1 + i
        conf   = ev.get("confidence_score", 0)
        clar   = ev.get("clarity_score", 0)
        pct    = (ev.get("logframe_linkage") or {}).get("pct_of_target")
        dir_ok = not (ev.get("logframe_linkage") or {}).get("direction_mismatch", False)
        values = [
            sub.get("indicator_name") or sub.get("result_statement", "")[:80],
            conf, clar,
            ev.get("verdict", ""),
            f"{pct:.0f}%" if pct is not None else "—",
            "Yes" if dir_ok else "No",
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font   = _font(bold=(col in (2, 3)), size=9)
            c.border = _border_thin()
            c.alignment = Alignment(wrap_text=True, vertical="top",
                                    horizontal="center" if col in (2, 3, 5, 6) else "left")
            if col == 2:
                c.fill = _score_fill(conf)
            elif col == 3:
                c.fill = _score_fill(clar)
            elif col == 6 and not dir_ok:
                c.fill = _fill(_RED_LIGHT)
        ws.row_dimensions[r].height = 32

    # Highest-leverage actions — Council XXVI: decision routing framing
    gap_row = start_row + n + 3
    ws.cell(row=gap_row, column=1,
            value="HIGHEST-LEVERAGE ACTIONS").font = _font(bold=True, colour=_RED_DARK, size=11)
    gap_row += 1

    # Compute average per dimension
    dim_keys = ["Directness", "Verification", "Recency", "Definition",
                "Measurement", "Integrity", "Scope", "Governance"]
    dim_maxes = {"Directness": 2.0, "Verification": 2.0, "Recency": 1.0,
                 "Definition": 1.25, "Measurement": 1.25, "Integrity": 1.0,
                 "Scope": 0.75, "Governance": 0.75}
    dim_comp_map = {
        "Directness":   ("confidence_components", "direct_score"),
        "Verification": ("confidence_components", "verify_score"),
        "Recency":      ("confidence_components", "recency_score"),
        "Definition":   ("clarity_components",    "definition_score"),
        "Measurement":  ("clarity_components",    "measurement_score"),
        "Integrity":    ("clarity_components",    "integrity_score"),
        "Scope":        ("clarity_components",    "scope_score"),
        "Governance":   ("clarity_components",    "governance_score"),
    }
    dim_avgs = {}
    for dim in dim_keys:
        comp_key, score_key = dim_comp_map[dim]
        scores = [ev.get(comp_key, {}).get(score_key, 0) for ev in evaluations]
        avg = sum(scores) / len(scores) if scores else 0
        dim_avgs[dim] = avg

    sorted_dims = sorted(dim_avgs.items(), key=lambda x: x[1] / dim_maxes[x[0]])
    hdr_lbl = ws.cell(row=gap_row, column=1, value="Sub-criterion")
    hdr_lbl.font = _font(bold=True, size=10)
    hdr_lbl.alignment = Alignment(wrap_text=True, vertical="top")
    hdr_avg = ws.cell(row=gap_row, column=2, value="Avg Score")
    hdr_avg.font = _font(bold=True, size=10)
    hdr_avg.alignment = Alignment(wrap_text=True, vertical="top")
    hdr_act = ws.cell(row=gap_row, column=3, value="System decision")
    hdr_act.font = _font(bold=True, size=10)
    hdr_act.alignment = Alignment(wrap_text=True, vertical="top")
    for dim, avg in sorted_dims:
        gap_row += 1
        pct_of_max = avg / dim_maxes[dim]
        if pct_of_max < 0.5:
            decision = "Fix this first — highest leverage across portfolio"
        elif pct_of_max < 0.75:
            decision = "Needs improvement — address after critical gaps"
        else:
            decision = "Acceptable — maintain current standard"
        lbl_c = ws.cell(row=gap_row, column=1, value=dim)
        lbl_c.font      = _font(size=9)
        lbl_c.alignment = Alignment(wrap_text=True, vertical="top")
        avg_c = ws.cell(row=gap_row, column=2, value=round(avg, 2))
        avg_c.fill      = _score_fill(avg * (5.0 / dim_maxes[dim]))
        avg_c.font      = _font(bold=True, size=9)
        avg_c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="top")
        dec_c = ws.cell(row=gap_row, column=3, value=decision)
        dec_c.font      = _font(size=9)
        dec_c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[gap_row].height = 20

    # Routing instruction at the bottom
    gap_row += 2
    ws.merge_cells(f"A{gap_row}:C{gap_row}")
    action_cell = ws.cell(row=gap_row, column=1,
        value=(
            "SYSTEM ROUTING: Start with sub-criteria marked 'Fix this first' — "
            "these are your portfolio's highest-leverage actions. "
            "Open Sheet 1 (Scored Results), sort by the weakest sub-criterion, "
            "and action the Fix 1 column for each affected result. "
            "Amber cells = auto-populated by AI — review before sharing with donors. "
            "Red cells = not found in document — fill manually."
        )
    )
    action_cell.font      = _font(size=9, bold=True, colour=_RED_DARK)
    action_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[gap_row].height = 60

    # How to re-score after fixes
    gap_row += 2
    ws.merge_cells(f"A{gap_row}:C{gap_row}")
    rs_title = ws.cell(row=gap_row, column=1, value="HOW TO RE-SCORE AFTER FIXES")
    rs_title.font      = _font(bold=True, colour=_GREEN_DARK, size=10)
    rs_title.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[gap_row].height = 18

    gap_row += 1
    ws.merge_cells(f"A{gap_row}:C{gap_row}")
    rs_cell = ws.cell(row=gap_row, column=1,
        value=(
            "PATH 1 — Fix source document → re-upload to Score My Report:\n"
            "Correct your Word or PDF report, then go to Portfolio Decision Audit → Score My Report tab "
            "and re-upload the updated document. ImpactProof will re-extract and re-score all results.\n\n"
            "PATH 2 — Edit Portfolio CSV → re-upload to CSV Portfolio tab (fastest for data fixes):\n"
            "From the Score My Report results page, click 'Download re-score CSV'. "
            "Open the CSV, correct the amber/red fields (e.g. evidence_description, verifier, evidence_date, "
            "logframe_target, logframe_achievement). Save as CSV. "
            "Then go to Portfolio Decision Audit → CSV Portfolio tab and upload the corrected file. "
            "ImpactProof re-scores all results and shows revised determinations, heatmap, and gap analysis.\n\n"
            "Column names in the CSV must stay unchanged for re-scoring to work."
        )
    )
    rs_cell.font      = _font(size=9)
    rs_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[gap_row].height = 110


# ---------------------------------------------------------------------------
# Re-score workbook — two sheets for the "Download re-score" button
# ---------------------------------------------------------------------------

_RESCORE_COL_WIDTHS = {
    "result_statement": 45, "target_group": 22, "timeframe": 14,
    "geographic_scope": 18, "evidence_type": 22, "evidence_description": 40,
    "evidence_date": 14, "internal_review": 22, "external_review": 22,
    "verifier": 25, "logframe_indicator": 30, "logframe_baseline": 16,
    "logframe_target": 16, "logframe_achievement": 18, "beneficiary_voice": 24,
    "additional_context": 25, "learning_notes": 22, "limitations_notes": 22,
    "qual_sourcing_documented": 18, "qual_triangulated": 16,
    "qual_bias_considered": 18, "qual_beneficiary_voice_represented": 22,
    "qual_consent_ethics_addressed": 22, "provenance_sampling": 20,
    "provenance_dedup": 20, "provenance_tool": 20, "provenance_independent": 20,
    "provenance_recall": 20, "provenance_traceable": 22,
}


def _determ_fill(conf: float | None, clar: float | None) -> PatternFill:
    lo = min(conf or 0, clar or 0)
    if lo >= 4.0:
        return _fill(_GREEN_LIGHT)
    if lo >= 2.5:
        return _fill(_AMBER_LIGHT)
    return _fill(_RED_LIGHT)


def _determ_label(conf: float | None, clar: float | None) -> str:
    lo = min(conf or 0, clar or 0)
    if lo >= 4.0:
        return "Submission-ready"
    if lo >= 2.5:
        return "Needs work"
    return "High risk"


def build_rescore_excel(
    rows: list[dict],
    evaluations: list[dict],
    statuses: list[dict] | None = None,
    org_name: str = "",
    doc_name: str = "",
) -> bytes:
    """Two-sheet re-score workbook.

    Sheet 1 'Re-score Data' — raw column names for re-upload to CSV Portfolio tab.
    Sheet 2 'Determinations' — colour-coded human-readable summary.
    """
    if not _OPENPYXL_OK:
        raise ImportError("openpyxl is required. Run: pip install openpyxl")

    wb = Workbook()

    # ── Sheet 1: Re-score Data ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Re-score Data"
    ws1.freeze_panes = "A2"

    col_keys = list(rows[0].keys()) if rows else list(_RESCORE_COL_WIDTHS.keys())

    for ci, key in enumerate(col_keys, start=1):
        hdr = ws1.cell(row=1, column=ci, value=key)
        hdr.fill      = _fill(_HEADER_BG)
        hdr.font      = _font(bold=True, colour=_HEADER_FG, size=9)
        hdr.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        hdr.border    = _border_thin()
        ws1.column_dimensions[get_column_letter(ci)].width = _RESCORE_COL_WIDTHS.get(key, 18)

    ws1.row_dimensions[1].height = 30

    # Add a tooltip comment on A1 so users know what to do with this sheet
    if _OPENPYXL_OK:
        note = Comment(
            "RE-SCORE DATA\n\n"
            "Edit cells in this sheet to correct amber/red fields from the Portfolio Audit Workbook.\n\n"
            "To re-score:\n"
            "1. Make corrections (e.g. evidence_description, verifier, provenance fields)\n"
            "2. Save this sheet as a CSV file (File → Save As → CSV)\n"
            "3. Upload the CSV to the CSV Portfolio tab in ImpactProof\n\n"
            "Column names must remain unchanged for re-scoring to work.",
            author="ImpactProof",
        )
        note.width = 360
        note.height = 200
        ws1.cell(row=1, column=1).comment = note

    for ri, row_dict in enumerate(rows, start=2):
        is_alt  = (ri % 2 == 0)
        base_bg = _GREY_LIGHT if is_alt else _WHITE
        for ci, key in enumerate(col_keys, start=1):
            val  = row_dict.get(key, "")
            cell = ws1.cell(row=ri, column=ci, value=str(val) if val else "")
            cell.font      = _font(size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border    = _border_thin()
            cell.fill      = _fill(base_bg)
        ws1.row_dimensions[ri].height = 36

    # Instruction footer
    foot_row = len(rows) + 3
    ws1.merge_cells(f"A{foot_row}:{get_column_letter(len(col_keys))}{foot_row}")
    fc = ws1.cell(row=foot_row, column=1,
                  value="HOW TO RE-SCORE: Edit fields above → File > Save As > CSV → "
                        "upload to CSV Portfolio tab in ImpactProof → revised determinations appear instantly.")
    fc.font      = _font(size=9, bold=True, colour=_AMBER_DARK)
    fc.alignment = Alignment(wrap_text=True, vertical="top")
    ws1.row_dimensions[foot_row].height = 36

    # ── Sheet 2: Determinations ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Determinations")
    ws2.freeze_panes = "A3"

    # Title
    ws2.merge_cells("A1:G1")
    title = ws2["A1"]
    ts    = datetime.now().strftime("%d %b %Y %H:%M")
    title.value = (
        f"ImpactProof — Portfolio Determinations"
        f"{' · ' + org_name if org_name else ''}"
        f"{' · ' + doc_name if doc_name else ''}"
        f" · {ts}"
    )
    title.fill      = _fill(_HEADER_BG)
    title.font      = _font(bold=True, colour=_HEADER_FG, size=12)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 28

    # Column headers
    det_cols = [
        ("#",                10),
        ("Result Statement", 45),
        ("Target Group",     22),
        ("Evidence Type",    22),
        ("Confidence",       12),
        ("Clarity",          12),
        ("Determination",    18),
        ("Fix — Priority 1", 40),
    ]
    for ci, (hdr_label, width) in enumerate(det_cols, start=1):
        hdr = ws2.cell(row=2, column=ci, value=hdr_label)
        hdr.fill      = _fill(_HEADER_BG)
        hdr.font      = _font(bold=True, colour=_HEADER_FG, size=10)
        hdr.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        hdr.border    = _border_thin()
        ws2.column_dimensions[get_column_letter(ci)].width = width
    ws2.row_dimensions[2].height = 24

    for ri, (row_dict, ev) in enumerate(zip(rows, evaluations), start=3):
        conf   = ev.get("confidence_score")
        clar   = ev.get("clarity_score")
        fixes  = ev.get("fixes", [])
        fix1   = fixes[0]["message"] if fixes else ""
        determ = _determ_label(conf, clar)
        dfill  = _determ_fill(conf, clar)
        is_alt = (ri % 2 == 0)
        base   = _GREY_LIGHT if is_alt else _WHITE

        row_vals = [
            (ri - 2,                               _fill(base),          "center"),
            (row_dict.get("result_statement", ""), _fill(base),          "left"),
            (row_dict.get("target_group",     ""), _fill(base),          "left"),
            (row_dict.get("evidence_type",    ""), _fill(base),          "left"),
            (conf,                                 _score_fill(conf),    "center"),
            (clar,                                 _score_fill(clar),    "center"),
            (determ,                               dfill,                "center"),
            (fix1,                                 _fill(base),          "left"),
        ]
        for ci, (val, fill, halign) in enumerate(row_vals, start=1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.fill      = fill
            cell.font      = _font(
                bold=(ci in (5, 6, 7)),
                colour=(_GREEN_DARK if determ == "Submission-ready" and ci == 7
                        else _RED_DARK if determ == "High risk" and ci == 7
                        else "000000"),
                size=9,
            )
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal=halign)
            cell.border    = _border_thin()
        ws2.row_dimensions[ri].height = 48

    # Legend
    leg = len(evaluations) + 4
    ws2.merge_cells(f"A{leg}:G{leg}")
    lc = ws2.cell(row=leg, column=1,
                  value="Determination thresholds: Submission-ready = both scores ≥ 4.0 | "
                        "Needs work = lowest score 2.5–3.9 | High risk = lowest score < 2.5")
    lc.font      = _font(size=8, colour="616161")
    lc.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[leg].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
